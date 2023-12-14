from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from django.views import View
from openpyxl import Workbook
from .forms import EstudianteForm, CursoForm, CalificacionesForm
from .models import Estudiante, Curso, Calificaciones


# Create your views here.
# views.py

def agregar_estudiante(request):
    if request.method == 'POST':
        form = EstudianteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_estudiantes')
    else:
        form = EstudianteForm()

    return render(request, 'agregar_estudiante.html', {'form': form})

def agregar_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_cursos')
    else:
        form = CursoForm()

    return render(request, 'agregar_curso.html', {'form': form})

def agregar_calificacion(request):
    if request.method == 'POST':
        form = CalificacionesForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_calificaciones')
    else:
        form = CalificacionesForm()

    return render(request, 'agregar_calificacion.html', {'form': form})

def lista_estudiantes(request):
    estudiantes = Estudiante.ver_todos()
    return render(request, 'lista_estudiantes.html', {'estudiantes': estudiantes})

def lista_cursos(request):
    cursos = Curso.ver_todos()
    return render(request, 'lista_cursos.html', {'cursos': cursos})

def lista_calificaciones(request):
    calificaciones = Calificaciones.ver_todas()
    return render(request, 'lista_calificaciones.html', {'calificaciones': calificaciones})

def modificar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(Estudiante, id=estudiante_id)

    if request.method == 'POST':
        form = EstudianteForm(request.POST, instance=estudiante)
        if form.is_valid():
            form.save()
            return redirect('lista_estudiantes')
    else:
        form = EstudianteForm(instance=estudiante)

    return render(request, 'modificar_estudiante.html', {'form': form, 'estudiante': estudiante})

def eliminar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(Estudiante, id=estudiante_id)

    if request.method == 'POST':
        estudiante.eliminar()
        return redirect('lista_estudiantes')

    return render(request, 'eliminar_estudiante.html', {'estudiante': estudiante})

def modificar_calificacion(request, calificacion_id):
    calificacion = get_object_or_404(Calificaciones, id=calificacion_id)

    if request.method == 'POST':
        form = CalificacionesForm(request.POST, instance=calificacion)
        if form.is_valid():
            form.save()
            return redirect('lista_calificaciones')
    else:
        form = CalificacionesForm(instance=calificacion)

    return render(request, 'modificar_calificacion.html', {'form': form, 'calificacion': calificacion})

def eliminar_calificacion(request, calificacion_id):
    calificacion = get_object_or_404(Calificaciones, id=calificacion_id)

    if request.method == 'POST':
        calificacion.eliminar()
        return redirect('lista_calificaciones')

    return render(request, 'eliminar_calificacion.html', {'calificacion': calificacion})

def modificar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            return redirect('lista_cursos')
    else:
        form = CursoForm(instance=curso)

    return render(request, 'modificar_curso.html', {'form': form, 'curso': curso})

def eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if request.method == 'POST':
        curso.eliminar()
        return redirect('lista_cursos')

    return render(request, 'eliminar_curso.html', {'curso': curso})

class ReporteEstudiantesExcel(View):
    def get(self, request, *args, **kwargs):
        estudiantes = Estudiante.ver_todos()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE ESTUDIANTES'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'Nombre'
        ws['C3'] = 'Apellido'
        ws['D3'] = 'Correo'
        cont = 4
        for estudiante in estudiantes:
            ws.cell(row=cont, column=2).value = estudiante.nombre
            ws.cell(row=cont, column=3).value = estudiante.apellido
            ws.cell(row=cont, column=4).value = estudiante.correo
            cont += 1
        nombre_archivo = "reporte_de_estudiantes.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

